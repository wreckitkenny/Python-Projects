#!/usr/bin/env python3
__author__ = "Wreck-it Kenny"
__copyright__ = "Copyright 2019, The Python Project"
__version__ = "1.1"
__email__ = "tung.tran.3295@gmail.com"
__status__ = "Production"
__doc__ = "A mini tool to get and add Amigo Staff information"

from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

# Object Class

## # # # # # # #
## Tkinter GUI #
## # # # # # # #
class initialPage():
    def __init__(self, master):
        self.master = master
        self.master.title(string='Amigo Staff Tool')
        self.master.resizable(width=False, height=False)
        self.master.iconbitmap(r'C:/Program Files/Amigo/Amigo Staff Tool/amigo.ico')
        self.frame_1()
        self.frame_2()
        self.frame_3()

    # Frame 1
    def frame_1(self):
        global filepath
        filepath = StringVar()
        frame = Frame(self.master, borderwidth=1, relief=FLAT)
        Entry(frame, textvariable=filepath).pack(side=LEFT, fill=BOTH, expand=YES, padx=3)
        Button(frame, text='Browse', width=8, command=browse).pack(side=RIGHT, padx=3)
        frame.pack(fill=X, padx=2, pady=5)

    # Frame 2
    def frame_2(self):
        global chosen
        chosen = IntVar()
        frame = Frame(self.master, borderwidth=1, relief=FLAT)
        Radiobutton(frame, text='Search a staff', value=1, variable=chosen).pack(side=LEFT, padx=30)
        Radiobutton(frame, text='Add a staff', value=2, variable=chosen).pack(side=LEFT, padx=30)
        frame.pack(fill=X, padx=2)

    # Frame 3
    def frame_3(self):
        global next_button
        frame = Frame(self.master, borderwidth=1, relief=FLAT)
        next_button = Button(frame, text='Next', width=8, state=DISABLED, command=choose)
        next_button.pack(side=RIGHT, padx=3)
        frame.pack(fill=X, padx=2, pady=5)

# # # # # # #
#  Window 2 #
# # # # # # #
class SearchPage():
    def __init__(self):
        self.master = Toplevel(root)
        self.master.geometry(newGeometry='300x300')
        self.master.title(string='Search an existing staff')
        self.master.iconbitmap(r'C:/Program Files/Amigo/Amigo Staff Tool/amigo.ico')
        self.master.resizable(width=False, height=False)
        self.search()
        self.result()

    # Frame: SEARCH FIELD
    def search(self):
        global search_name
        search_name = StringVar()
        search_frame = Frame(self.master, borderwidth=1)

        search_frame2 = Frame(search_frame, relief=GROOVE, borderwidth=2)
        Entry(search_frame2, width=35, textvariable=search_name).pack(side=LEFT, fill=BOTH, expand=YES, pady=10, padx=5)
        Button(search_frame2, text='Search', width=8, command=get_information).pack(side=RIGHT, fill=X, padx=5)
        search_frame2.pack(fill=BOTH, padx=3, pady=10)

        Label(search_frame, text='Name to search:').place(relx=0.06, rely=0.15, anchor=W)
        search_frame.pack(fill=BOTH)

    # Frame: SEARCH RESULT
    def result(self):
        global result_info
        result_info = StringVar()
        result_frame = Frame(self.master)

        result_frame2 = Frame(result_frame, relief=GROOVE, borderwidth=2)
        Message(result_frame2, textvariable=result_info).pack(fill=BOTH, expand=YES)
        result_frame2.pack(fill=BOTH, expand=YES, padx=4, pady=5)

        Label(result_frame, text='Staff Information:').place(relx=0.06, rely=0.015, anchor=W)
        result_frame.pack(fill=BOTH, expand=YES)

# # # # # # #
#  Window 3 #
# # # # # # #
class AddPage():
    def __init__(self):
        self.master = Toplevel(root)
        self.master.title(string='Add a new staff')
        self.master.iconbitmap(r'C:/Program Files/Amigo/Amigo Staff Tool/amigo.ico')
        self.master.resizable(width=False, height=False)
        self.add()

    def add(self):
        Label(self.master, text='Name:').grid(row=0, sticky=W, padx=5, pady=5)
        Label(self.master, text='Title:').grid(row=1, sticky=W, padx=5, pady=5)
        Label(self.master, text='Phone:').grid(row=2, sticky=W, padx=5, pady=5)
        Label(self.master, text='Email:').grid(row=3, sticky=W, padx=5, pady=5)
        global add_name, add_title, add_phone, add_email
        add_name = StringVar()
        add_title = StringVar()
        add_phone = StringVar()
        add_email = StringVar()
        Entry(self.master, width=25, textvariable=add_name).grid(row=0, column=1, padx=5, pady=5)
        Entry(self.master, width=25, textvariable=add_title).grid(row=1, column=1, padx=5, pady=5)
        Entry(self.master, width=25, textvariable=add_phone).grid(row=2, column=1, padx=5, pady=5)
        Entry(self.master, width=25, textvariable=add_email).grid(row=3, column=1, padx=5, pady=5)
        Button(self.master, text='Add', width=8, command=add_information).grid(row=0, column=2, padx=5, pady=5)
        Button(self.master, text='Cancel', width=8, command=self.master.quit).grid(row=1, column=2, padx=5)


class Done():
    def __init__(self):
        self.done = Tk()
        self.done.resizable(width=False, height=False)
        self.end()
        self.done.mainloop()

    def backtomain(self):
        self.done.destroy()
        main()

    def end(self):
        Label(self.done, text='Added successfully!!!').pack(fill=BOTH, expand=YES, padx=10, pady=10)
        done_frame = Frame(self.done)
        Button(done_frame, width=10, text='Back to main', command=self.backtomain).pack(side=LEFT, padx=5)
        Button(done_frame, width=10, text='Exit', command=self.done.quit).pack(side=LEFT, padx=5)
        done_frame.pack(fill=BOTH, expand=YES, padx=20, pady=20)

## App Functions
def get_information():
    for i in range(len(name)):
        if name[i].value == ' '.join([i.capitalize() for i in search_name.get().rstrip().split(' ')]):
            global infor
            infor = """======================\n
|| Name: {}\n
|| Title: {}\n
|| Phone: {}\n
|| Email: {}\n
======================""".format(name[i].value, title[i].value, phone[i].value, email[i].value)
            result_info.set(infor)


def add_information():
        add_num = num[::-1][0].value + 1
        ws.append((add_num, add_name.get(), add_title.get(), '', add_phone.get(), add_email.get()))
        for col in ws.columns:
            aligned_cell = col[-1]
            aligned_cell.alignment = Alignment(vertical='center')
            aligned_cell.font = Font(name='Times New Roman', size=12)
            aligned_cell.border = Border(left=Side(border_style='thin'),
                                         right=Side(border_style='thin'),
                                         top=Side(border_style='thin'),
                                         bottom=Side(border_style='thin'))
        wb.save(filename)
        initialPage(root).master.destroy()
        Done()


def browse():
    global filename
    filename = str(filedialog.askopenfilename(
        initialdir = "C:\\",
        title = "Select Excel file",
        filetypes = (("EXCEL files","*.xlsx"),("all files","*.*"))))
    filepath.set(filename)
    if filename != '':
        next_button['state'] = NORMAL
    
    # Initial definitions
    global wb, ws, num, name, title, phone, email
    wb = load_workbook(filename)
    ws = wb['AMG']
    num = ws['A']
    name = ws['B']
    title = ws['C']
    phone = ws['E']
    email = ws['F']


def choose():
    if chosen.get() == 1 :
        SearchPage()

    if chosen.get() == 2:
        AddPage()


def main():
    global root
    root = Tk()
    initial_page = initialPage(root)
    root.mainloop()


if __name__ == "__main__":
    main()