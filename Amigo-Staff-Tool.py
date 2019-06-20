#!/usr/bin/env python3
__author__ = "Wreck-it Kenny"
__copyright__ = "Copyright 2019, The Python Project"
__version__ = "1.0"
__email__ = "tung.tran.3295@gmail.com"
__status__ = "Production"
__doc__ = "A mini tool to get and add Amigo Staff information using OPENPYXL"

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

location = "\\\\10.0.8.5\\Public IS\\Danh sach nhan vien Amigo\\Staff list in\
 Amigo June2019.xlsx"
wb = openpyxl.load_workbook(location)
ws = wb['AMG']
num = ws['A']
name = ws['B']
title = ws['C']
phone = ws['E']
email = ws['F']


class Amigo_Staff():
    def __init__(self, staff_name):
        self.staff_name = staff_name

    def get_information(self):
        for i in range(len(name)):
            if name[i].value == self.staff_name:
                print(f"""==============================
Name: {name[i].value}
Title: {title[i].value}
Phone: {phone[i].value}
Email: {email[i].value}
==============================""")
                break
        else:
            print("Not Found")

    def add_information(self, staff_title, staff_phone, staff_email):
        self.staff_num = num[::-1][0].value + 1
        ws.append((self.staff_num, self.staff_name, staff_title, '',
                   staff_phone, staff_email))
        for col in ws.columns:
            aligned_cell = col[-1]
            aligned_cell.alignment = Alignment(horizontal='center',
                                               vertical='center')
            aligned_cell.font = Font(name='Times New Roman', size=12)
            aligned_cell.border = Border(left=Side(border_style='thin'),
                                         right=Side(border_style='thin'),
                                         top=Side(border_style='thin'),
                                         bottom=Side(border_style='thin'),
                                         diagonal=Side(border_style='thin'),
                                         outline=Side(border_style='thin'),
                                         vertical=Side(border_style='thin'),
                                         horizontal=Side(border_style='thin'))
        wb.save(location)


if __name__ == '__main__':
    while True:
        print("""
What do you want?
1. Search a staff information
2. Add a new staff""")
        res = int(input("Choose: "))
        if res == 1:
            s = input("What name to search? ")
            ppl = Amigo_Staff(s)
            ppl.get_information()
        elif res == 2:
            try:
                new_name = input("What is name? ")
                new_title = input("What is title? ")
                new_phone = int(input("What is phone number? "))
                new_email = input("What is email? ")
                ppl = Amigo_Staff(new_name)
                ppl.add_information(staff_title=new_title,
                                    staff_phone=new_phone,
                                    staff_email=new_email)
                print("Successfully added a new staff")
            except PermissionError:
                print("""
==================WARNING!=================
Please close the Excel file before updating""")
            break
        else:
            print("{} is an invalid option".format(res))
